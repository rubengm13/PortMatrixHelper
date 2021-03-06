import openpyxl
import time
from pathlib import Path
import optparse
import sys
import os
import Network.Network as Network

# Import TextFSM
os.environ["NET_TEXTFSM"] = str(Path(os.getcwd())/Path("Network/ntc-templates/templates"))

DEFAULT_IGNORE_SHEETS = ["Comments", "Settings"]

VERBOSE = False


def cli_args():
    """Reads the CLI options provided and returns them using the OptionParser
    Will return the Values as a dictionary"""
    parser = optparse.OptionParser()
    parser.add_option('-v','--verbose',
                      dest="verbose",
                      default=False,
                      action="store_true",
                      help="Enable Verbose Output."
                      )
    parser.add_option('-i','--input_file',
                      dest="input_file",
                      default="PortMatrix.xlsx",
                      action="store",
                      help="Input file name of excel sheet."
                      )
    parser.add_option('-o','--output',
                      dest="output_file",
                      action="store",
                      help="Output file name of excel sheet, if not output file is specified it will overwrite the input file."
                      )
    parser.add_option('-g','--generate_config',
                      dest="generate_config",
                      default=False,
                      action="store_true",
                      help="Generate Configuration based on the matching template."
                      )
    parser.add_option('-c','--check_connections',
                      dest="check_connections",
                      default=False,
                      action="store_true",
                      help="Logs in to the devices, checks if connection matches the outline connection in the spreadsheet. Will check via CDP and LLDP. Will also gather SN information and other basic info from 'show version'."
                      )
    options, remainder = parser.parse_args()
    # Utilizing the vars() method we can return the options as a dictionary
    return vars(options)


def open_xls(xls_file_name):
    """Returns the WorkBook of specified Name
    Name of XLS file must be imported as a Path"""
    if not xls_file_name.exists():
        print("The following file does not exists:", xls_file_name)
        print("Please ensure the file exists or the correct filename was entered when utilizing the '-i | --input_file' option.")
        sys.exit()
    if VERBOSE:
        print("Opening Excel sheet:", xls_file_name)
    return openpyxl.load_workbook(xls_file_name, data_only=True)


def add_xls_tag(file_name):
    """Check the file_name to ensure it has ".xlsx" extension, if not add it"""
    if file_name[:-5] != ".xlsx":
        return file_name +".xlsx"
    else:
        return file_name


def rw_cell(sheet_obj, row, column, value=None):
    """Either writes or reads to/from a cell.
    To Erase a cell input the value as an empty string: ''."""
    if value != None:
        sheet_obj.cell(row=row, column=column).value = value
        return sheet_obj.cell(row=row, column=column).value
    return sheet_obj.cell(row=row, column=column).value


def get_config_templates(ws_obj):
    """Reads the Templates from the Excel Sheet"""
    return_dict = {}
    for row in range(1, 1+len(ws_obj["A"])):
        key = rw_cell(ws_obj, row, 1)
        value = rw_cell(ws_obj, row, 2)
        if key != None:
            if key not in return_dict.keys():
                return_dict[key] = value
            else:
                print("A template by the Name '%s' already exists, please ensure all template Names are unique."%key)
                print("Exiting Now.")
                sys.exit()
    return return_dict


def get_ignore_sheets(ws_obj):
    """Get the List of Sheet Names to Ignore"""
    return_list = DEFAULT_IGNORE_SHEETS.copy()
    for row in range(2, 1+len(ws_obj["D"])):
        value = rw_cell(ws_obj, row, 4)
        if value:
            return_list.append(value)
    return return_list


def gen_cfg_by_ws(wb_obj, config_templates, ignore_sheets, header_index):
    """Cycles through the Worksheets and will ignore any worksheets in the
    ignore_sheets list as defined in 'Settings' Sheet"""
    for sheetname  in wb_obj.sheetnames:
        if sheetname not in ignore_sheets:
            gen_config_to_cell(wb_obj[sheetname], config_templates, header_index)


def gen_config_to_cell(ws_obj, config_templates, header_col_index):
    """Utilize the a dictionary mapping key to a configuration template, cfg template
    should be in str.format() method form for a dictionary.
    Will ignore all columns before the header_col_index value.
    Issues with adding the Configuration at the End of the Columns, A "Configuration
    """
    # Need the Header Name to Index Value Dictionary to identify locations of Headers
    header_index = get_table_headers(ws_obj, True, header_col_index)
    if "Configuration" not in header_index.keys():
        print("Issue with this: ", ws_obj,
                "\nThis Sheet needs to have a header named 'Configuration' defined in Row",
                header_col_index, "\nThis is to ensure that the output is printed correctly.")
        return
    for row in ws_obj:
        row_index = row[0].row
        if row_index > header_col_index:
            if row[header_index["Template"]-1].value:
                # Ned the Row as a Dictionary to easily utilize the str.format() method
                row_dict=gen_row_dict(row, header_index)
                tmplt = config_templates[row[header_index["Template"]-1].value]
                cfg = tmplt.format(**row_dict)
                rw_cell(ws_obj, row_index, header_index["Configuration"], cfg)


def get_table_headers(ws_obj, dict, index):
    """returns the Table Headers, if dict is true it will return a dictionary with the index as the value and the header as the key
    If you change the Column that it returns as the 'Headers'"""
    if dict:
        return_dict = {}
        for cell in ws_obj[index]:
            return_dict[cell.value] = cell.col_idx
        return return_dict
    else:
        return_list
        for cell in ws_obj[index]:
            return_list.append(cell)
        return return_list


def gen_row_dict(row, header_index):
    """Takes a Row of cells and a header_index to generate a dictionary for that row"""
    return_dict = {}
    for header,col_index in header_index.items():
        cell_value = row[col_index-1].value
        if cell_value:
            return_dict[header]=cell_value
        else:
            return_dict[header]=""
    return return_dict


def save_xls(wb_obj, file_name, out_dir_path=Path("")):
    """Saves the WorkBook to provided Directory and File Name."""
    file_save_string = out_dir_path
    file_save_string = file_save_string/file_name
    print("saving the file to:", file_save_string)
    wb_obj.save(file_save_string)


def read_device_information(wb_obj, ignore_sheets):
    net_devices = []
    for sheetname in wb_obj.sheetnames:
        if sheetname not in ignore_sheets:
            net_dev_info = {
                "host":rw_cell(wb_obj[sheetname], 1, 2),
                "username":rw_cell(wb_obj[sheetname], 2, 2),
                "password":rw_cell(wb_obj[sheetname], 3, 2),
                "secret":rw_cell(wb_obj[sheetname], 4, 2),
                "device_type":rw_cell(wb_obj[sheetname], 5, 2),
                "sheetname":sheetname
            }
            if None not in net_dev_info.values():
                net_devices.append(Network.NetworkDevice(**net_dev_info))
    return net_devices


def check_net_dev_connection(net_dev, wb_obj, header_index):
    ws_obj = wb_obj[net_dev.sheetname]
    for row in range(header_index+1, ws_obj.max_row+1):
        neigh_info={
            "local_interface":rw_cell(ws_obj, row, 1),
            "neighbor":rw_cell(ws_obj, row, 2),
            "remote_interface":rw_cell(ws_obj, row, 3)
        }
        dict_values = neigh_info.values()
        if None not in dict_values and "" not in dict_values:
            connection_status=""
            response = net_dev.verify_cdp_neigh(**neigh_info)
            if response:
                connection_status += response+"\n"
            response = net_dev.verify_lldp_neigh(**neigh_info)
            #print(response)
            if response:
                connection_status += response
            if connection_status:
                rw_cell(ws_obj, row, 4, connection_status)


def update_discovered_data(net_dev, wb_obj):
    ws_obj = wb_obj[net_dev.sheetname]
    if net_dev.status == "Complete":
        rw_cell(ws_obj, 1, 4, net_dev.hostname)
        rw_cell(ws_obj, 2, 4, net_dev.version)
        if isinstance(net_dev.model, list):
            rw_cell(ws_obj, 3, 4, ",".join(net_dev.model))
        else:
            rw_cell(ws_obj, 3, 4, net_dev.model)
        if isinstance(net_dev.serial_number, list):
            rw_cell(ws_obj, 4, 4, ",".join(net_dev.serial_number))
        else:
            rw_cell(ws_obj, 4, 4, net_dev.serial_number)
        rw_cell(ws_obj, 5, 4, net_dev.boot_image)
    rw_cell(ws_obj, 6, 2, net_dev.status)


def check_all_devices_connections(wb_obj, ignore_sheets, header_index):
    # Gatheres the Devices and connects to them and logs all the data from them.
    net_devices = read_device_information(wb_obj, ignore_sheets)
    #print(net_devices)
    for net_dev in net_devices:
        if net_dev.status == "Data Gathered":
            interfaces = check_net_dev_connection(net_dev, wb_obj, header_index)
            net_dev.status = "Complete"
        update_discovered_data(net_dev, wb_obj)


###### MAIN ######
def main():
    header_index = 9
    # Load some cli Args and adjust necessary VERBOSE Option
    setup_args = cli_args()
    if setup_args["verbose"] == True:
        global VERBOSE
        VERBOSE = True
        Network.VERBOSE = True

    wb_obj = open_xls(Path(setup_args["input_file"]))

    config_templates = get_config_templates(wb_obj["Settings"])
    ignore_sheets = get_ignore_sheets(wb_obj["Settings"])
    if setup_args["check_connections"]:
        check_all_devices_connections(wb_obj, ignore_sheets, header_index)

    if setup_args["generate_config"]:
        gen_cfg_by_ws(wb_obj, config_templates,ignore_sheets, header_index)
    # Save Configuration
    if setup_args["output_file"]:
        setup_args["output_file"] = add_xls_tag(setup_args["output_file"])
        save_xls(wb_obj, setup_args["output_file"])
    else:
        save_xls(wb_obj, setup_args["input_file"])


if __name__ == "__main__":
    main()
